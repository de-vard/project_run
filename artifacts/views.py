from openpyxl import load_workbook
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.views import APIView

from artifacts.models import CollectibleItem
from artifacts.serializers import CollectibleItemSerializer


class CollectibleItemViewSet(viewsets.ModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


class UploadExcelData(APIView):
    """Загрузка и парсинг данных"""

    def post(self, request):
        file = request.FILES.get('file')

        if not file:
            return Response(
                {'detail': 'Файл не передан'},
                status=status.HTTP_400_BAD_REQUEST
            )

        wb = load_workbook(file)
        sheet = wb.active

        invalid_rows = []

        # пропускаем заголовок
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = {
                'name': row[0],
                'uid': row[1],
                'value': row[2],
                'latitude': row[3],
                'longitude': row[4],
                'picture': row[5],
            }

            serializer = CollectibleItemSerializer(data=data)

            if serializer.is_valid():
                serializer.save()
            else:

                invalid_rows.append(list(row))

        #  возвращаем СПИСОК, а не dict
        return Response(invalid_rows, status=status.HTTP_200_OK)
